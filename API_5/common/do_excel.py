from openpyxl import load_workbook
from API_5.common import project_path  #导入测试用例的绝对路径
from API_5.common.read_config import ReadConfig
class DoExcsl:
    """完成Excel读写的操作"""
    def __init__(self,file_name,sheet_name):
       self.file_name = file_name
       self.sheet_name = sheet_name
       self.case_conf =ReadConfig(project_path.conf_path).get_str("My_log","case_id")  #获取日志信息
       self.write = ReadConfig(project_path.conf_path).get_str("My_log","write")  #表示写入的表单

    def get_init_data(self):   #获取初始化的手机号
        wb = load_workbook(project_path.case_path) #打开工作簿
        sheet =wb["sheet1"]                        #定位表单
        moblie = sheet.cell(1,2).value             #读取指定位置的手机号这里获取的是整数
        return moblie                              #返回获取的手机号
    def update_init_data(self,value):
        wb = load_workbook(project_path.case_path) #打开工作簿
        sheet = wb["sheet1"]                       #定位表单
        sheet.cell(1,2).value=value                #指定位置更新手机号
        wb.save(project_path.case_path)  #保存工作簿
        wb.close()                      #关闭工作簿
    def login_excel(self):  #读取充值模块的用例
        # case_id = ReadConfig(project_path11.conf_path).get_data("My_log", "case_id")
        # print(case_id)
        wb = load_workbook(project_path.case_path)  # 打开表格
        sheet = wb["login"]
        # 储存所有的数据 # 注意这里获取的是整数
        text_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data["Case_id"] = sheet.cell(i, 1).value
            sub_data["Module"] = sheet.cell(i,2).value
            sub_data["url"] = sheet.cell(i, 3).value
            sub_data["Title"] = sheet.cell(i, 4).value
            sub_data["Method"] = sheet.cell(i, 5).value
            sub_data["Params"] = sheet.cell(i,6).value
            sub_data["ExpectedResult"] = sheet.cell(i,8).value
            text_data.append(sub_data)
        final_data = []
        if self.case_conf == "all":  # 配置文件控制用例如果是列表，那就获取列表里面指定id对应的数据
            final_data = text_data  # 把测试用例赋值给final_data这个变量
        else:
            for i in self.case_conf:  # 遍历case_id里面的值
                final_data.append(text_data[i-1])
        return text_data
    def recharge_test(self):  #读取充值模块的用例，提现的用例
        # case_id = ReadConfig(project_path11.conf_path).get_data("My_log", "case_id")
        # print(case_id)
        wb = load_workbook(project_path.case_path)  # 打开表格
        sheet = wb[self.write]
        # 储存所有的数据 # 注意这里获取的是整数
        text_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data["Case_id"] = sheet.cell(i, 1).value
            sub_data["Module"] = sheet.cell(i,2).value
            sub_data["url"] = sheet.cell(i, 3).value
            sub_data["Title"] = sheet.cell(i, 4).value
            sub_data["Method"] = sheet.cell(i, 5).value
            sub_data["Params"] = sheet.cell(i,6).value
            sub_data["Sql"] = sheet.cell(i, 7).value
            sub_data["ExpectedResult"] = sheet.cell(i, 8).value
            text_data.append(sub_data)
        final_data = []
        if self.case_conf == "all":  # 配置文件控制用例如果是列表，那就获取列表里面指定id对应的数据
            final_data = text_data  # 把测试用例赋值给final_data这个变量
        else:
            for i in self.case_conf:  # 遍历case_id里面的值
                final_data.append(text_data[i-1])
        return text_data
    def add_excel(self):  #读取充值模块的用例
        # case_id = ReadConfig(project_path11.conf_path).get_data("My_log", "case_id")
        # print(case_id)
        wb = load_workbook(project_path.case_path)  # 打开表格
        sheet = wb["add"]
        # 储存所有的数据 # 注意这里获取的是整数
        text_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data["Case_id"] = sheet.cell(i, 1).value
            sub_data["Module"] = sheet.cell(i, 2).value
            sub_data["url"] = sheet.cell(i, 3).value
            sub_data["Title"] = sheet.cell(i, 4).value
            sub_data["Method"] = sheet.cell(i, 5).value
            sub_data["Params"] = sheet.cell(i, 6).value
            sub_data["Sql"] = sheet.cell(i, 7).value
            sub_data["ExpectedResult"] = sheet.cell(i, 8).value
            text_data.append(sub_data)
        final_data = []
        if self.case_conf == "all":  # 配置文件控制用例如果是列表，那就获取列表里面指定id对应的数据
            final_data = text_data  # 把测试用例赋值给final_data这个变量
        else:
            for i in self.case_conf:  # 遍历case_id里面的值
                final_data.append(text_data[i-1])
        return text_data
    def write_excel(self, row, column, values):
        #row代表写入的行
        #column代表写入的列
        #values代表写入值
            '''在指定的单元格写入指定的数据，并保存到当前Excel'''
            try:  # 输入的变量可能导致异常， 例如文件名、表单名不存在，因此加入异常处理
                wb = load_workbook(project_path.case_path)
                sheet = wb[self.write]
                sheet.cell(row, column).value = values
                wb.save(project_path.case_path)
            except KeyError as e:
                print("您的输入有误，报错为：{}".format(e))  # 打印报错信息
                exit()  # 退出程序，注意这里不能用break，break用于for和while循环中


if __name__ == '__main__':
    text_data = DoExcsl(project_path.case_path,"recharge").recharge_test()
    print(text_data)


